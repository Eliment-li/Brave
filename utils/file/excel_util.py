
import pandas as pd
import  re
import openpyxl
class ExcelUtil:
    @staticmethod
    def read_by_sheet(excel_file):
        # Load the Excel file
        xls = pd.ExcelFile(excel_file)

        # Get a list of all sheet names
        sheet_names = xls.sheet_names

        # Print the list of sheet names
        #print(sheet_names)

        data = {}
        for sheet_name in sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            data.update({sheet_name:df})
        #print(data['qnn'])
        return  sheet_names,data

    @staticmethod
    #2d array to a sheet
    def array2sheet(file_name, sheet_name, data):
        try:
            # Try to load the existing workbook
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            # If the file does not exist, create a new workbook
            workbook = openpyxl.Workbook()

        # If the sheet does not exist, create it
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
        else:
            sheet = workbook[sheet_name]

        # Write the 2D array to the sheet
        for row_index, row in enumerate(data):
            for col_index, value in enumerate(row):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=value)

        # Save the workbook
        workbook.save(file_name)

    # Example usage
    data = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9]
    ]

# read excel data to np array
    @staticmethod
    def read_sheet_to_array(file_name, sheet_name='Sheet2'):
        try:
            # Load the Excel file
            df = pd.read_excel(file_name, sheet_name=sheet_name,header=None)
            # Convert DataFrame to numpy array
            return df.to_numpy()
        except FileNotFoundError:
            print(f"File {file_name} not found.")
            return None
        except ValueError as e:
            print(f"Error reading sheet {sheet_name}: {e}")
            return None

if __name__ == '__main__':
    # Example usage
    # data = [
    #     [1, 2, 3],
    #     [4, 5, 6],
    #     [7, 8, 9]
    # ]
    #
    # ExcelUtil.array2sheet('d:/example.xlsx', 's.a', data)
    array = ExcelUtil.read_sheet_to_array('d:/layout.xlsx')
    print(repr(array))